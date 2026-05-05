import os
import subprocess
import datetime
import pandas as pd
import openpyxl
from openpyxl.cell.cell import MergedCell
import csv
import time
import xlwings as xw
from ftplib import FTP

WORK_DIR = '/Users/christianvidalwolf/Stock'
MASTER_FILE = f'{WORK_DIR}/INICIO PLUS 2023.xlsx'
SKUS_FORZAR_CERO = ['2450'] # ID for ASIN B010TN6SXU
PRECIOS_FIJOS = {
    "1016VCI": 10.99,
    "14165VCI": 8.99,
    "40552MD": 22.99,
    "42772MD": 26.99,
    "4359VCI": 10.99,
    "31469MD": 13.99,
    "2388VCI": 15.99,
    "2185651CLM": 9.99,
    "23779SG": 15.99,
    "17385VCI": 8.99,
    "31181MD": 21.99,
    "1304VC": 10.99,
    "15846VCI": 12.99,
    "11302VC": 16.99,
}

SIGNES_SKUS_FORZAR_CERO = {
    '11631VC', '1237VC', '1238VC', '1652VC', '1653VC', '1684VC', '1688VC', '1717VC',
    '1718VC', '180VC', '1832VC', '2025VC', '2027VC', '2180VC', '2181VC', '2210VC',
    '2260VC', '2355VC', '2356VC', '2360VC', '2400VC', '2401VC', '2402VC', '2405VC',
    '2406VC', '2414VC', '2415VC', '2425VC', '2447VC', '2510VC', '2517VC', '2522VC',
    '2611VC', '2648VC', '2660VC', '2665VC', '2857VC', '2858VC', '2990VC', '3011VC',
    '3012VC', '3013VC', '3015VC', '3018VC', '3019VC', '3020VC', '3021VC', '3027VC',
    '3030VC', '3035VC', '3038VC', '3041VC', '3043VC', '3044VC', '3045VC', '3046VC',
    '3048VC', '3052VC', '3053VC', '3054VC', '3055VC', '3062VC', '3063VC', '3064VC',
    '3069VC', '3070VC', '3071VC', '3072VC', '3074VC', '3089VC', '3116VC', '3117VC',
    '3118VC', '3119VC', '3120VC', '3121VC', '3123VC', '3125VC', '3126VC', '3127VC',
    '3130VC', '3133VC', '3134VC', '3135VC', '3200VC', '3201VC', '3230VC', '3311VC',
    '3314VC', '3315VC', '3354VC', '3372VC', '3380VC', '3391VC', '3392VC', '3423VC',
    '3432VC', '3437VC', '3484VC', '3487VC', '3555VC', '3557VC', '3558VC', '3559VC',
    '3712VC', '3824VC', '4107VC', '4108VC', '4110VC', '4115VC', '4125VC', '4130VC',
    '4134VC', '4140VC', '4155VC', '4158VC', '4165VC', '4231VC', '4232VC', '4255VC',
    '4259VC', '4261VC', '4263VC', '4264VC', '4265VC', '4285VC', '4288VC', '4330VC',
    '4361VC', '4442VC', '4445VC', '4462VC', '4528VC', '4531VC', '4534VC', '4558VC',
    '4657VC', '4721VC', '4722VC', '4727VC', '4735VC', '4737VC', '4745VC', '4752VC',
    '4758VC', '4765VC', '4785VC', '4790VC', '4801VC', '4815VC', '4874VC', '4882VC',
    "2915244CLM", "30512912CLM", "2861262CLM", "41683MDRG", "41891MDRG",
    '5032VC', '5113VC', '5382VC', '5383VC', '5440VC', '5442VC', '5456VC', '5540VC',
    '5571VC', '5572VC', '5574VC', '5575VC', '5588VC', '5592VC', '5685VC', '5712VC',
    '5721VC', '5807VC', '5823VC', '5825VC', '5835VC', '5852VC', '5871VC', '5872VC',
    '5873VC', '5874VC', '5882VC', '5883VC', '5885VC', '5987VC', '5991VC', '5995VC',
    '6001VC', '6007VC', '6040VC', '6151VC', '6252VC', '6255VC', '6270VC', '6471VC',
    '5658VCI',
}

ASINS_FORZAR_CERO = {
    'B0GM8HT8VG',
    'B0GM119SWH',
    'B076NB2QK7',
    'B0F1NGM1KK',
    'B0F1NJ2ZP9',
    'B0G1J29M7N',
    'B01D0MP66C',
    'B01D0MQBF2',
    'B01D0MQ6CU',
    'B01D0MQKSK',
    'B01D0MQ5UI',
    'B01D0MPCNE',
    'B01D0MQOJU',
    'B07D9T4G1F',
    'B07D9V2R3J',
}

def to_num(val):
    if val is None or val == "":
        return 0
    try:
        val_str = str(val).replace(',', '.')
        if '.' in val_str:
            return float(val_str)
        return int(float(val_str))
    except:
        return val

def download_files():
    dcasa_file = ""
    minerales_file = f'{WORK_DIR}/minerales_feed.xml'
    signes_file = f'{WORK_DIR}/signes_stock.csv'
    
    # Download DCASA
    print("Downloading DCASA...")
    for attempt in range(1, 7):
        try:
            print(f"Attempting to connect to FTP data.dcasacollection.com (attempt {attempt}/6)...")
            ftp = FTP('data.dcasacollection.com', timeout=180)
            ftp.set_pasv(True)
            ftp.login(user='sek4283', passwd='Rx34z5m6_pER')
            files = ftp.nlst()
            data_files = [f for f in files if f.startswith('DataWeb') and f.endswith('.csv')]
            if data_files:
                latest_file = sorted(data_files)[-1]
                dcasa_file = f'{WORK_DIR}/{latest_file}'
                print(f"Downloading {latest_file} from FTP...")
                with open(dcasa_file, 'wb') as fp:
                    ftp.retrbinary(f'RETR {latest_file}', fp.write)
            ftp.quit()
            break # success
        except Exception as e:
            print(f"Error downloading from DCASA (attempt {attempt}/6): {e}")
            if attempt < 6:
                wait_time = 15 * attempt
                print(f"Waiting {wait_time}s before next attempt...")
                time.sleep(wait_time)
            else:
                print("Max attempts reached for DCASA download. Attempting to use local fallback...")

    # Fallback to local if download failed
    if not dcasa_file or not os.path.exists(dcasa_file):
        local_files = sorted([f for f in os.listdir(WORK_DIR) if f.startswith('DataWeb') and f.endswith('.csv')])
        for fname in reversed(local_files):
            fpath = f"{WORK_DIR}/{fname}"
            # Check if file has enough rows
            try:
                with open(fpath, encoding='latin-1') as chk:
                    rc = sum(1 for _ in chk) - 1
                if rc >= 1000:
                    dcasa_file = fpath
                    print(f"Using fallback local DCASA file: {fname}")
                    break
            except: continue

    # Download MINERALES
    print("Downloading MINERALES...")
    subprocess.run(['curl', '-L', 'https://vivescortadaimport.com/modules/doofinder/feed2.php?language=ES&currency=EUR', '-o', minerales_file])
    
    # Download SIGNES
    print("Downloading SIGNES...")
    subprocess.run(['curl', '-L', 'https://signesconexion.com/stock/STOCK-44880.CSV', '-o', signes_file])
    
    return dcasa_file, minerales_file, signes_file

# --- DCASA ---
def process_dcasa(input_file):
    if not input_file or not os.path.exists(input_file):
        print("DCASA input file not found. Skipping...")
        return
    
    print(f"--- Processing DCASA: {input_file} ---")

    # Validación: archivo incompleto si tiene menos de 1000 filas
    with open(input_file, mode='r', encoding='latin-1') as f:
        row_count = sum(1 for _ in f) - 1  # excluir cabecera
    if row_count < 1000:
        print(f"WARNING: DCASA file has only {row_count} rows (expected >1000). Skipping to avoid overwriting with incomplete data.")
        return

    def clean_float(val):
        if not val: return 0.0
        return float(val.replace(',', '.'))

    cleaned_rows = []
    with open(input_file, mode='r', encoding='latin-1') as f:
        reader = csv.reader(f, delimiter=';')
        header = next(reader)
        for row in reader:
            if len(row) < 18: continue
            try:
                tarifa_a = clean_float(row[4])
                stock = clean_float(row[17])
                if stock <= 3: continue
                if stock < 20 and tarifa_a < 20: continue
                cleaned_rows.append(row)
            except ValueError:
                continue

    wb_master = openpyxl.load_workbook(MASTER_FILE)
    ws_dcasa = wb_master['InicioDcasa']
    start_row = 148
    
    max_row = ws_dcasa.max_row
    if max_row >= start_row:
        for r in range(start_row, max_row + 1):
            for c in range(1, 10):
                cell = ws_dcasa.cell(row=r, column=c)
                if not isinstance(cell, MergedCell):
                    cell.value = None

    for r_idx, row in enumerate(cleaned_rows, start=start_row):
        cols = [
            to_num(row[1]), to_num(row[1]), row[2], 
            to_num(row[17]), to_num(row[4]), row[13], 
            row[16], to_num(row[12]), to_num(row[7])
        ]
        for c_idx, val in enumerate(cols, start=1):
            cell = ws_dcasa.cell(row=r_idx, column=c_idx)
            if not isinstance(cell, MergedCell):
                cell.value = val
                
    # Build stock lookup from cleaned_rows for DcasaWeb update
    dcasa_stock_map = {}
    for row in cleaned_rows:
        code = to_num(row[1])
        stock = to_num(row[17])
        if code:
            dcasa_stock_map[code] = stock

    # Update DcasaWeb col E directly to avoid Excel smart-recalculation missing new rows
    ws_dcasaweb = wb_master['DcasaWeb']
    updated = 0
    for r in range(2, ws_dcasaweb.max_row + 1):
        code_cell = ws_dcasaweb.cell(row=r, column=2)
        if code_cell.value is None:
            continue
        code = to_num(code_cell.value)
        stock = dcasa_stock_map.get(code, 0)
        if stock == 0:
            val = 0
        elif stock == 1:
            val = 1
        else:
            val = 99
        ws_dcasaweb.cell(row=r, column=5).value = val
        updated += 1

    wb_master.save(MASTER_FILE)
    wb_master.close()
    print(f"DCASA update finished. Rows added: {len(cleaned_rows)}, DcasaWeb rows updated: {updated}")

# --- MINERALES ---
def process_minerales(input_file):
    print(f"--- Processing MINERALES: {input_file} ---")
    df = pd.read_csv(input_file, sep='|', encoding='utf-8')
    def clean_num(val):
        try: return float(val) if not pd.isna(val) else 0.0
        except: return 0.0

    # Discard products with stock < 5, regardless of price (User request)
    df = df[df['stock'].apply(clean_num) >= 5]
    
    wb_master = openpyxl.load_workbook(MASTER_FILE)
    ws_min = wb_master['StockMIN']
    
    max_row = ws_min.max_row
    if max_row >= 2:
        for r in range(2, max_row + 1):
            for c in range(1, 5): 
                cell = ws_min.cell(row=r, column=c)
                if not isinstance(cell, MergedCell):
                    cell.value = None

    for r_idx, row in enumerate(df.itertuples(index=False), start=2):
        l_id_val = str(row[0]).replace('.0', '')
        # If it's one of the SKUs to force to 0, or stock < 5, it already was filtered 
        # but let's be explicit if we wanted to keep the row but set stock 0.
        # However, here we just follow the mapping logic.
        stock_val = to_num(row[7])
        if l_id_val in SKUS_FORZAR_CERO:
            stock_val = 0

        price_val = to_num(row[12])
        if str(row[0]) in PRECIOS_FIJOS:
            price_val = PRECIOS_FIJOS[str(row[0])]

        row_vals = [to_num(row[0]), stock_val, stock_val, price_val]
        for c_idx, val in enumerate(row_vals, start=1):
            cell = ws_min.cell(row=r_idx, column=c_idx)
            if not isinstance(cell, MergedCell):
                cell.value = val
                
    wb_master.save(MASTER_FILE)
    wb_master.close()
    print(f"MINERALES update finished. Rows added: {len(df)}")

# --- SIGNES ---
def process_signes(input_file):
    print(f"--- Processing SIGNES: {input_file} ---")
    def clean_stock(val):
        try: return float(str(val).replace(',', '.')) if not (not val or val == "") else 0.0
        except: return 0.0

    def clean_price(val):
        try: return float(str(val).replace(',', '.')) if not (not val or val == "") else 0.0
        except: return 0.0
            
    def remove_sg(val):
        return str(val).replace('SG-', '')

    cleaned_rows = []
    with open(input_file, mode='r', encoding='latin-1') as f:
        reader = csv.reader(f, delimiter=';')
        for row in reader:
            if not row: continue
            stock = clean_stock(row[2])
            price = clean_price(row[3])
            
            # Force 0 for '83627' (User request)
            if '83627' in str(row[0]):
                stock = 0

            # Force 0 for VC SKUs (User request)
            raw_sku = str(row[0]).strip()
            if '2088' in raw_sku:
                print(f"[DEBUG 2088] raw_sku={repr(raw_sku)}, stock={stock}")
            if any(sku in raw_sku.upper() for sku in [s.upper() for s in SIGNES_SKUS_FORZAR_CERO]):
                stock = 0

            if stock <= 3: continue
            if stock < 20 and price < 20: continue
            cleaned_rows.append(row)

    wb_master = openpyxl.load_workbook(MASTER_FILE)
    ws_signes = wb_master['InicioSignes']
    start_row = 188
    
    max_row = ws_signes.max_row
    if max_row >= start_row:
        for r in range(start_row, max_row + 1):
            for c in range(1, 9):
                cell = ws_signes.cell(row=r, column=c)
                if not isinstance(cell, MergedCell):
                    cell.value = None

    for r_idx, row in enumerate(cleaned_rows, start=start_row):
        cols = [
            to_num(remove_sg(row[0])), row[1], "", 
            to_num(row[3]), to_num(row[6]), to_num(row[7]), 
            to_num(row[8]), to_num(row[2])
        ]
        for c_idx, val in enumerate(cols, start=1):
            cell = ws_signes.cell(row=r_idx, column=c_idx)
            if not isinstance(cell, MergedCell):
                cell.value = val

    wb_master.save(MASTER_FILE)
    wb_master.close()
    print(f"SIGNES update finished. Rows added: {len(cleaned_rows)}")


def export_amazon():
    print("--- Exporting to STOCK AMZ.txt ---")
    target_txt = f'{WORK_DIR}/STOCK AMZ.txt'
    temp_calc_file = f'{WORK_DIR}/temp_calculated.xlsx'
    
    app = None
    try:
        print("Starting Excel to bake in formula values...")
        app = xw.App(visible=True)
        app.display_alerts = False
        wb = app.books.open(MASTER_FILE)
        
        print("Calculating...")
        app.calculate()
        time.sleep(30)
        
        print(f"Saving a temporary calculated copy to {temp_calc_file}...")
        if os.path.exists(temp_calc_file):
            os.remove(temp_calc_file)
        wb.save(temp_calc_file)
        wb.close()
        app.quit()
        app = None
        
        print("Reading values from calculated copy using openpyxl...")
        wb_data = openpyxl.load_workbook(temp_calc_file, data_only=True)
        ws = wb_data['STOCK ES']
        
        with open(target_txt, 'w', encoding='utf-8', newline='') as f:
            writer = csv.writer(f, delimiter='\t')
            for r_idx, row in enumerate(ws.iter_rows(values_only=True)):
                if any(x is not None for x in row):
                    # Force stock=0 for ASINs in ASINS_FORZAR_CERO (col 0 = ASIN, col 4 = stock)
                    asin = str(row[0]).strip() if row[0] is not None else ''
                    if asin in ["2915244CLM", "30512912CLM", "2861262CLM", "41683MDRG", "41891MDRG"]:
                        continue
                    row = list(row)
                    force_zero = asin in ASINS_FORZAR_CERO or asin.upper() in {s.upper() for s in SIGNES_SKUS_FORZAR_CERO}
                    
                    # Override price if in PRECIOS_FIJOS
                    if asin in PRECIOS_FIJOS:
                        row[1] = PRECIOS_FIJOS[asin]
                        # Update min/max prices accordingly
                        row[2] = round(row[1] / 2, 2)
                        row[3] = round(row[1] * 2, 2)

                    if '2088' in asin:
                        print(f"[DEBUG EXPORT] asin={repr(asin)}, stock_before={row[4]}, force_zero={force_zero}")
                    if force_zero:
                        row[4] = 0
                    clean_row = []
                    for col_idx, val in enumerate(row):
                        # Empty cols 5 and 6
                        if r_idx > 0 and col_idx in [5, 6]:
                            clean_row.append("")
                            continue
                            
                        if val is None:
                            clean_row.append("")
                        elif isinstance(val, (int, float)):
                            if col_idx in [1, 2, 3]:
                                clean_row.append(f"{round(float(val), 2):g}".replace('.', ','))
                            elif col_idx == 4:
                                clean_row.append(str(int(round(float(val)))))
                            else:
                                if float(val).is_integer():
                                    clean_row.append(str(int(val)))
                                else:
                                    clean_row.append(f"{round(float(val), 4):g}".replace('.', ','))
                        else:
                            clean_row.append(str(val))
                    writer.writerow(clean_row)
        
        if os.path.exists(temp_calc_file):
            os.remove(temp_calc_file)
        print(f"Export finished successfully. File updated: {target_txt}")

    except Exception as e:
        print(f"FATAL ERROR in export: {e}")
    finally:
        if app:
            try: app.quit()
            except: pass
        if os.path.exists(temp_calc_file):
            try: os.remove(temp_calc_file)
            except: pass

if __name__ == "__main__":
    print(f"--- Starting Stock Update @ {datetime.datetime.now()} ---")
    dcasa_file, minerales_file, signes_file = download_files()
    if dcasa_file: process_dcasa(dcasa_file)
    process_minerales(minerales_file)
    process_signes(signes_file)
    export_amazon()
    print(f"--- Finished Stock Update @ {datetime.datetime.now()} ---")
